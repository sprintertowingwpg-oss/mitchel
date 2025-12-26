#!/usr/bin/env python3
"""Plot grouped_by_truck.csv 'total without discount' as a bar chart.

Generates `grouped_totals.png` in the workspace root. Adds a final bar 'All vehicles'
which is the sum of all `total without discount` values.

Usage: python3 scripts/plot_grouped.py grouped_by_truck.csv
"""
import csv
import sys
import argparse
from pathlib import Path
import matplotlib.pyplot as plt
import math
from datetime import datetime


def detect_date_range_from_invoices():
    p = Path('invoices.csv')
    if not p.exists():
        return None
    try:
        with open(p, newline='') as fh:
            r = csv.DictReader(fh)
            headers = r.fieldnames or []
            date_hdr = None
            for h in headers:
                if h and h.strip().lower() in ('date', 'posted on', 'posted_on', 'postedon'):
                    date_hdr = h
                    break
            if not date_hdr:
                return None
            dates = []
            for row in r:
                v = (row.get(date_hdr) or '').strip()
                if not v:
                    continue
                for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
                    try:
                        d = datetime.strptime(v, fmt)
                        dates.append(d)
                        break
                    except Exception:
                        continue
            if not dates:
                return None
            return f"{min(dates).strftime('%Y-%m-%d')} — {max(dates).strftime('%Y-%m-%d')}"
    except Exception:
        return None


def detect_customer_from_invoices():
    p = Path('invoices.csv')
    if not p.exists():
        return None
    try:
        with open(p, newline='') as fh:
            r = csv.DictReader(fh)
            if 'customer' not in (r.fieldnames or []):
                return None
            vals = set()
            for row in r:
                v = (row.get('customer') or '').strip()
                if v:
                    vals.add(v)
                if len(vals) > 1:
                    return None
            return next(iter(vals)) if vals else None
    except Exception:
        return None


def read_grouped(path):
    rows = []
    path = str(path)
    if path.lower().endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SystemExit('openpyxl is required to read XLSX files.')
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # Support both vehicle and owner grouping
        if 'vehicle' in headers:
            label_idx = headers.index('vehicle')
            title = 'vehicle'
        elif 'owner' in headers:
            label_idx = headers.index('owner')
            title = 'owner'
        else:
            raise SystemExit('Could not find required columns in XLSX.')
        total_idx = headers.index('Total') if 'Total' in headers else None
        if total_idx is None:
            raise SystemExit('Could not find Total column in XLSX.')
        for row in ws.iter_rows(min_row=2):
            label = str(row[label_idx].value).strip() if row[label_idx].value else 'Unknown'
            try:
                total = float(row[total_idx].value or 0)
            except Exception:
                total = 0.0
            rows.append((label, total))
    else:
        with open(path, encoding='utf-8') as f:
            r = csv.DictReader(f)
            for line in r:
                vehicle = (line.get('vehicle') or '').strip() or 'Unknown'
                try:
                    total = float((line.get('Total') or '0').replace(',', ''))
                except Exception:
                    total = 0.0
                rows.append((vehicle, total))
    return rows


def human(x):
    if x >= 1e9:
        return f'{x/1e9:.2f}B'
    if x >= 1e6:
        return f'{x/1e6:.2f}M'
    if x >= 1e3:
        return f'{x/1e3:.2f}K'
    return f'{x:.2f}'


def plot(rows, out_png='grouped_totals.png', subtitle=None):
    # sort descending for better visualization
    rows_sorted = sorted(rows, key=lambda x: x[1], reverse=True)
    labels = [r[0] for r in rows_sorted]
    values = [r[1] for r in rows_sorted]
    total_all = sum(values)

    # do not add an extra bar for the total; instead display it in the title/subtitle

    fig, ax = plt.subplots(figsize=(max(8, len(labels)*0.6), 6))
    bars = ax.bar(labels, values, color='tab:blue')

    ax.set_ylabel('Total')
    title = 'Total per vehicle'
    total_text = f'Total (all vehicles): {human(total_all)}'
    if subtitle:
        ax.set_title(title + '\n' + subtitle + ' — ' + total_text)
    else:
        ax.set_title(title + '\n' + total_text)
    # set ticks and labels explicitly to avoid Matplotlib warnings
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=45, ha='right')
    # leave room for rotated labels
    fig.subplots_adjust(bottom=0.30)

    # annotate bars
    for bar, val in zip(bars, values):
        h = bar.get_height()
        ax.annotate(human(val), xy=(bar.get_x() + bar.get_width() / 2, h), xytext=(0, 3),
                    textcoords='offset points', ha='center', va='bottom', fontsize=8)

    plt.tight_layout()
    plt.savefig(out_png, dpi=150)
    print(f'Wrote chart: {out_png}')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('csv', nargs='?', default='grouped_by_truck.csv')
    parser.add_argument('--out', default='grouped_totals.png')
    parser.add_argument('--customer', help='Customer name to show on chart')
    parser.add_argument('--date-range', help='Explicit date range text to show on chart')
    args = parser.parse_args()

    p = Path(args.csv)
    if not p.exists():
        print('Grouped CSV not found:', p)
        sys.exit(1)
    rows = read_grouped(p)
    subtitle_parts = []
    if args.customer:
        subtitle_parts.append(args.customer)
    if args.date_range:
        subtitle_parts.append(args.date_range)
    if not subtitle_parts:
        dr = detect_date_range_from_invoices()
        if dr:
            subtitle_parts.append(dr)
    subtitle = ' — '.join(subtitle_parts) if subtitle_parts else None
    plot(rows, out_png=args.out, subtitle=subtitle)


if __name__ == '__main__':
    main()
#!/usr/bin/env python3
"""Plot grouped by truck totals.

Reads `grouped_by_truck.csv`, uses the `total without discount` per truck for bars,
computes overall total without taxes (per-truck `total without discount` minus `total of taxes`),
and saves a bar chart to `grouped_by_truck_total_wo_taxes.png`.
"""
import csv
from pathlib import Path
import math


def read_grouped(csv_path):
    rows = []
    with open(csv_path, encoding='utf-8') as f:
        r = csv.DictReader(f)
        for line in r:
            rows.append(line)
    return rows


def tofloat(s):
    try:
        return float((s or '').replace(',', '').strip())
    except Exception:
        return 0.0


def plot(rows, out_png):
    try:
        import matplotlib.pyplot as plt
    except Exception:
        raise

    # Build data
    trucks = []
    values = []
    taxes = []
    for r in rows:
        truck = r.get('truck') or r.get('truck ') or 'Unknown'
        tot_wo_discount = tofloat(r.get('total without discount') or r.get('total_without_discount') or r.get('total without discount '))
        tot_taxes = tofloat(r.get('total of taxes') or r.get('total_of_taxes') or r.get('total of taxes '))
        trucks.append(truck)
        values.append(tot_wo_discount)
        taxes.append(tot_taxes)

    # compute overall total without taxes
    total_wo_taxes = sum(v - t for v, t in zip(values, taxes))

    # sort by value desc for nicer plot
    combined = sorted(zip(trucks, values), key=lambda x: x[1], reverse=True)
    trucks_sorted = [c[0] for c in combined]
    values_sorted = [c[1] for c in combined]

    plt.figure(figsize=(max(8, len(trucks_sorted)*0.6), 6))
    bars = plt.bar(range(len(values_sorted)), values_sorted, color='tab:blue')
    plt.xticks(range(len(trucks_sorted)), trucks_sorted, rotation=45, ha='right')
    plt.ylabel('Total without discount (currency)')
    plt.title('Total without discount per truck\nTotal without taxes (all vehicles): {:,.2f}'.format(total_wo_taxes))
    plt.tight_layout()
    # annotate bars
    for b, v in zip(bars, values_sorted):
        h = b.get_height()
        plt.text(b.get_x()+b.get_width()/2, h, f'{v:,.0f}', ha='center', va='bottom', fontsize=8)

    plt.savefig(out_png)
    print('Wrote chart:', out_png)


if __name__ == '__main__':
    # Only use the argparse-based entry point
    import sys
    if len(sys.argv) > 1:
        # If arguments are provided, use the argparse-based main
        pass  # argparse-based main is already defined above
    else:
        print("Please provide the grouped CSV file as an argument, e.g.:\n  python3 scripts/plot_grouped.py grouped_by_truck.csv --out grouped_totals.png")
