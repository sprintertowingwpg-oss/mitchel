#!/usr/bin/env python3
"""
Generate all reports and charts from a Crystal Reports XML input.

Usage:
    python3 scripts/generate_all_reports.py test.xml

This will produce:
    - invoices.csv, invoices.xlsx
    - grouped_by_truck.csv, grouped_by_truck.xlsx
    - grouped_totals.png (bar chart)
    - labor_vs_parts_pie.png (pie chart)
"""

import sys
import subprocess
from pathlib import Path
import shutil
import csv
import re

def run(cmd, desc=None):
    print(f"\n[Running] {desc or ' '.join(cmd)}")
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(result.stdout)
        print(result.stderr)
        raise SystemExit(f"Failed: {' '.join(cmd)}")
    print(result.stdout)

def get_last_invoice_date(xlsx_path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl is required to read XLSX files.")
        return None
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    date_col = None
    for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1))):
        if str(cell.value).strip().lower() == 'date':
            date_col = idx
            break
    if date_col is None:
        return None
    dates = []
    date_re = re.compile(r'\d{1,2}/\d{1,2}/\d{2,4}')
    for row in ws.iter_rows(min_row=2):
        v = row[date_col].value
        if v and isinstance(v, str) and date_re.match(v.strip()):
            dates.append(v.strip())
    if not dates:
        return None
    # Parse dates and return the latest
    from datetime import datetime
    dt_objs = []
    for d in dates:
        for fmt in ('%m/%d/%Y', '%Y-%m-%d'):
            try:
                dt_objs.append(datetime.strptime(d, fmt))
                break
            except Exception:
                continue
    if not dt_objs:
        return None
    return max(dt_objs).strftime('%Y-%m-%d')

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/generate_all_reports.py <input.xml>")
        sys.exit(1)
    xml = sys.argv[1]
    # 1. Extract all invoices to XLSX (for all data)
    temp_xlsx = 'invoices.xlsx'
    temp_group_xlsx = 'grouped_by_truck.xlsx'
    run([
        sys.executable, 'scripts/extract_invoices.py', xml,
        '--group',
        '--xlsx', temp_xlsx,
        '--group-xlsx', temp_group_xlsx,
    ], desc='Extracting invoices and grouped report (XLSX only)')

    # 2. Find last date in invoices.xlsx for main output folder
    last_date = get_last_invoice_date(temp_xlsx)
    if not last_date:
        last_date = 'unknown_date'
    outdir = Path(last_date)
    outdir.mkdir(exist_ok=True)

    # 3. Move XLSX files to output folder (before plotting)
    out_invoices_xlsx = outdir / 'invoices.xlsx'
    out_grouped_xlsx = outdir / 'grouped_by_truck.xlsx'
    shutil.move(temp_xlsx, out_invoices_xlsx)
    shutil.move(temp_group_xlsx, out_grouped_xlsx)

    # Generate grouped_by_owner.xlsx and chart for all data
    import importlib.util
    spec = importlib.util.spec_from_file_location("extract_invoices", str(Path('scripts/extract_invoices.py')))
    ei = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(ei)
    # Read all invoices
    from openpyxl import load_workbook
    wb_all = load_workbook(out_invoices_xlsx, read_only=True)
    ws_all = wb_all.active
    headers_all = [str(cell.value).strip() if cell.value else '' for cell in next(ws_all.iter_rows(min_row=1, max_row=1))]
    rows_all = []
    for row in ws_all.iter_rows(min_row=2):
        rows_all.append({h: v for h, v in zip(headers_all, [cell.value for cell in row])})
    grouped_owner = ei.group_by_owner(rows_all)
    out_grouped_owner_xlsx = outdir / 'grouped_by_owner.xlsx'
    ei.write_group_owner_xlsx(grouped_owner, out_grouped_owner_xlsx)
    run([
        sys.executable, 'scripts/plot_grouped.py', str(out_grouped_owner_xlsx), '--out', str(outdir / 'grouped_owners.png')
    ], desc='Generating bar chart (grouped_owners.png)')
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/generate_all_reports.py <input.xml>")
        sys.exit(1)
    xml = sys.argv[1]
    # 1. Extract all invoices to XLSX (for all data)
    temp_xlsx = 'invoices.xlsx'
    temp_group_xlsx = 'grouped_by_truck.xlsx'
    run([
        sys.executable, 'scripts/extract_invoices.py', xml,
        '--group',
        '--xlsx', temp_xlsx,
        '--group-xlsx', temp_group_xlsx,
    ], desc='Extracting invoices and grouped report (XLSX only)')

    # 2. Find last date in invoices.xlsx for main output folder
    last_date = get_last_invoice_date(temp_xlsx)
    if not last_date:
        last_date = 'unknown_date'
    outdir = Path(last_date)
    outdir.mkdir(exist_ok=True)

    # 3. Move XLSX files to output folder (before plotting)
    out_invoices_xlsx = outdir / 'invoices.xlsx'
    out_grouped_xlsx = outdir / 'grouped_by_truck.xlsx'
    shutil.move(temp_xlsx, out_invoices_xlsx)
    shutil.move(temp_group_xlsx, out_grouped_xlsx)

    # 4. Generate bar chart (total per vehicle)
    run([
        sys.executable, 'scripts/plot_grouped.py', str(out_grouped_xlsx), '--out', str(outdir / 'grouped_totals.png')
    ], desc='Generating bar chart (grouped_totals.png)')

    # 5. Generate pie chart (labor vs parts)
    run([
        sys.executable, 'scripts/plot_pie_labor_parts.py', str(out_grouped_xlsx), '--out', str(outdir / 'labor_vs_parts_pie.png')
    ], desc='Generating pie chart (labor_vs_parts_pie.png)')

    # 6. Generate reports for each month
    print("\nGenerating reports for each month...")
    from openpyxl import load_workbook, Workbook
    from collections import defaultdict
    import calendar
    wb = load_workbook(out_invoices_xlsx, read_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    date_idx = headers.index('Date') if 'Date' in headers else None
    if date_idx is None:
        print('No Date column found in invoices.xlsx, skipping monthly reports.')
        return
    # Collect rows by year-month
    rows_by_month = defaultdict(list)
    for row in ws.iter_rows(min_row=2):
        date_val = row[date_idx].value
        if not date_val:
            continue
        # Try to parse date
        from datetime import datetime
        dt = None
        for fmt in ('%m/%d/%Y', '%Y-%m-%d'):
            try:
                dt = datetime.strptime(str(date_val), fmt)
                break
            except Exception:
                continue
        if not dt:
            continue
        ym = dt.strftime('%Y-%m')
        rows_by_month[ym].append([cell.value for cell in row])

    for ym, rows in sorted(rows_by_month.items()):
        month_dir = outdir / ym
        month_dir.mkdir(exist_ok=True)
        # Write invoices_month.xlsx
        month_xlsx = month_dir / 'invoices.xlsx'
        wb_month = Workbook()
        ws_month = wb_month.active
        ws_month.append(headers)
        for r in rows:
            ws_month.append(r)
        wb_month.save(month_xlsx)
        # Group by vehicle for this month
        invoice_dicts = []
        for r in rows:
            d = {h: v for h, v in zip(headers, r)}
            invoice_dicts.append(d)
        grouped = ei.group_by_vehicle(invoice_dicts)
        # Write grouped_by_truck.xlsx for this month
        month_group_xlsx = month_dir / 'grouped_by_truck.xlsx'
        ei.write_group_xlsx(grouped, month_group_xlsx)
        # Group by owner for this month
        grouped_owner = ei.group_by_owner(invoice_dicts)
        month_group_owner_xlsx = month_dir / 'grouped_by_owner.xlsx'
        ei.write_group_owner_xlsx(grouped_owner, month_group_owner_xlsx)
        # Generate charts for this month
        run([
            sys.executable, 'scripts/plot_grouped.py', str(month_group_xlsx), '--out', str(month_dir / 'grouped_totals.png')
        ], desc=f'Generating bar chart for {ym}')
        run([
            sys.executable, 'scripts/plot_pie_labor_parts.py', str(month_group_xlsx), '--out', str(month_dir / 'labor_vs_parts_pie.png')
        ], desc=f'Generating pie chart for {ym}')
        run([
            sys.executable, 'scripts/plot_grouped.py', str(month_group_owner_xlsx), '--out', str(month_dir / 'grouped_owners.png')
        ], desc=f'Generating bar chart for owners in {ym}')
        print(f"Reports and charts for {ym} generated in {month_dir}")

    # 7. Generate reports for each quarter
    print("\nGenerating reports for each quarter...")
    def get_quarter(dt):
        return f"{dt.year}-Q{((dt.month-1)//3)+1}"
    rows_by_quarter = defaultdict(list)
    for row in ws.iter_rows(min_row=2):
        date_val = row[date_idx].value
        if not date_val:
            continue
        from datetime import datetime
        dt = None
        for fmt in ('%m/%d/%Y', '%Y-%m-%d'):
            try:
                dt = datetime.strptime(str(date_val), fmt)
                break
            except Exception:
                continue
        if not dt:
            continue
        qtr = get_quarter(dt)
        rows_by_quarter[qtr].append([cell.value for cell in row])

    for qtr, rows in sorted(rows_by_quarter.items()):
        qtr_dir = outdir / qtr
        qtr_dir.mkdir(exist_ok=True)
        # Write invoices.xlsx for the quarter
        qtr_xlsx = qtr_dir / 'invoices.xlsx'
        wb_qtr = Workbook()
        ws_qtr = wb_qtr.active
        ws_qtr.append(headers)
        for r in rows:
            ws_qtr.append(r)
        wb_qtr.save(qtr_xlsx)
        # Group by vehicle for this quarter
        invoice_dicts = []
        for r in rows:
            d = {h: v for h, v in zip(headers, r)}
            invoice_dicts.append(d)
        grouped = ei.group_by_vehicle(invoice_dicts)
        # Write grouped_by_truck.xlsx for this quarter
        qtr_group_xlsx = qtr_dir / 'grouped_by_truck.xlsx'
        ei.write_group_xlsx(grouped, qtr_group_xlsx)
        # Group by owner for this quarter
        grouped_owner = ei.group_by_owner(invoice_dicts)
        qtr_group_owner_xlsx = qtr_dir / 'grouped_by_owner.xlsx'
        ei.write_group_owner_xlsx(grouped_owner, qtr_group_owner_xlsx)
        # Generate charts for this quarter
        run([
            sys.executable, 'scripts/plot_grouped.py', str(qtr_group_xlsx), '--out', str(qtr_dir / 'grouped_totals.png')
        ], desc=f'Generating bar chart for {qtr}')
        run([
            sys.executable, 'scripts/plot_pie_labor_parts.py', str(qtr_group_xlsx), '--out', str(qtr_dir / 'labor_vs_parts_pie.png')
        ], desc=f'Generating pie chart for {qtr}')
        run([
            sys.executable, 'scripts/plot_grouped.py', str(qtr_group_owner_xlsx), '--out', str(qtr_dir / 'grouped_owners.png')
        ], desc=f'Generating bar chart for owners in {qtr}')
        print(f"Reports and charts for {qtr} generated in {qtr_dir}")

    print(f"\nAll reports and charts generated in folder: {outdir}\nAnd per-month and per-quarter reports in subfolders.")

if __name__ == '__main__':
    main()