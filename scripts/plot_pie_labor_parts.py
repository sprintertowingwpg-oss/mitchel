#!/usr/bin/env python3
"""Generate a pie chart comparing total labor vs total parts.

Usage: python3 scripts/plot_pie_labor_parts.py grouped_by_truck.csv --out labor_vs_parts_pie.png
"""
import argparse
import csv
import math
import os
from pathlib import Path
from datetime import datetime
import matplotlib.pyplot as plt


def parse_num(s):
    if s is None:
        return 0.0
    s = str(s).strip()
    if s == "":
        return 0.0
    # remove currency symbols and commas
    s = s.replace('$', '').replace('€', '').replace(',', '')
    # handle parentheses for negative
    neg = False
    if s.startswith('(') and s.endswith(')'):
        neg = True
        s = s[1:-1]
    try:
        v = float(s)
    except Exception:
        # fallback: keep digits and dot and minus
        filtered = ''.join(ch for ch in s if ch.isdigit() or ch in '.-')
        v = float(filtered) if filtered not in ('', '.', '-', '-.') else 0.0
    return -v if neg else v


def find_header(headers, candidates):
    lower = [h.lower() for h in headers]
    for c in candidates:
        if c.lower() in lower:
            return headers[lower.index(c.lower())]
    # try contains
    for h in headers:
        for c in candidates:
            if c.lower() in h.lower():
                return h
    return None


def main():
    p = argparse.ArgumentParser()
    p.add_argument('csv', help='Grouped CSV (or invoices CSV)')
    p.add_argument('--out', default='labor_vs_parts_pie.png', help='Output image')
    p.add_argument('--customer', help='Customer name to show on chart')
    p.add_argument('--date-range', help='Date range text to show on chart')
    args = p.parse_args()

    labor_sum = 0.0
    parts_sum = 0.0
    input_path = str(args.csv)
    if input_path.lower().endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SystemExit('openpyxl is required to read XLSX files.')
        wb = load_workbook(input_path, read_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        labor_idx = headers.index('Labor') if 'Labor' in headers else None
        parts_idx = headers.index('Parts') if 'Parts' in headers else None
        if labor_idx is None or parts_idx is None:
            raise SystemExit('Could not find required columns in XLSX.')
        for row in ws.iter_rows(min_row=2):
            labor_sum += parse_num(row[labor_idx].value)
            parts_sum += parse_num(row[parts_idx].value)
    else:
        with open(input_path, newline='') as fh:
            reader = csv.DictReader(fh)
            for r in reader:
                labor_sum += parse_num(r.get('Labor'))
                parts_sum += parse_num(r.get('Parts'))

    total = labor_sum + parts_sum
    if math.isclose(total, 0.0):
        raise SystemExit('No labor or parts amounts found (sums are zero). Check CSV headers.')

    labels = ['Labor', 'Parts']
    sizes = [labor_sum, parts_sum]

    fig, ax = plt.subplots(figsize=(6, 6))
    wedges, texts, autotexts = ax.pie(
        sizes,
        labels=labels,
        autopct=lambda pct: f"{pct:.1f}%\n({int(round(pct/100*total)):,})",
        startangle=90,
        colors=['#4c78a8', '#f58518'],
    )
    ax.axis('equal')
    # build subtitle: prefer explicit args, else try to detect from invoices.csv
    subtitle_parts = []
    if args.customer:
        subtitle_parts.append(args.customer)
    if args.date_range:
        subtitle_parts.append(args.date_range)
    if not subtitle_parts:
        # try detect date range from invoices.csv in workspace
        inv_path = Path('invoices.csv')
        if inv_path.exists():
            try:
                with open(inv_path, newline='') as fh:
                    rdr = csv.DictReader(fh)
                    date_hdr = None
                    headers = rdr.fieldnames or []
                    for h in headers:
                        if h and h.strip().lower() in ('date', 'posted on', 'posted_on', 'postedon'):
                            date_hdr = h
                            break
                    if date_hdr:
                        dates = []
                        for r in rdr:
                            v = (r.get(date_hdr) or '').strip()
                            for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
                                try:
                                    d = datetime.strptime(v, fmt)
                                    dates.append(d)
                                    break
                                except Exception:
                                    continue
                        if dates:
                            lo = min(dates).strftime('%Y-%m-%d')
                            hi = max(dates).strftime('%Y-%m-%d')
                            subtitle_parts.append(f'{lo} — {hi}')
            except Exception:
                pass

    # try detect single customer from invoices.csv if not provided
    if not args.customer:
        inv_path = Path('invoices.csv')
        if inv_path.exists():
            try:
                with open(inv_path, newline='') as fh:
                    rdr = csv.DictReader(fh)
                    if 'customer' in (rdr.fieldnames or []):
                        customers = set()
                        for r in rdr:
                            v = (r.get('customer') or '').strip()
                            if v:
                                customers.add(v)
                            if len(customers) > 1:
                                break
                        if len(customers) == 1:
                            args.customer = next(iter(customers))
            except Exception:
                pass

    if args.customer:
        # prefer to show customer first in subtitle
        if subtitle:
            subtitle = args.customer + ' — ' + subtitle
        else:
            subtitle = args.customer

    subtitle = ' — '.join(subtitle_parts) if subtitle_parts else None
    if subtitle:
        ax.set_title('Total labor vs total parts\n' + subtitle)
    else:
        ax.set_title('Total labor vs total parts')

    plt.tight_layout()
    fig.savefig(args.out, dpi=150)
    print('Wrote chart:', args.out)


if __name__ == '__main__':
    main()
